#!/usr/bin/env python3
"""
Prozesskostenrechner — Excel-Tool für Handelsgerichtsprozesse ZH/AG
Basierend auf «From Tariff to Award», Model 6
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


# ================================================================
# FARBPALETTE
# ================================================================
C_NAVY       = "1B2A4A"
C_MED_NAVY   = "2C3E6B"
C_BLUE       = "3574A7"
C_WHITE      = "FFFFFF"
C_TEXT        = "333333"
C_HINT       = "8B8B8B"
C_INPUT_BG   = "FFF8E1"
C_INPUT_BD   = "D4A843"
C_OUTPUT_BG  = "EBF1EB"
C_TOTAL_BG   = "C8E6C9"
C_RANGE_BG   = "FDF2E6"
C_RANGE_BD   = "E8C58A"
C_DISCL_BG   = "FFF3CD"
C_DISCL_BD   = "C9A825"
C_DISCL_TXT  = "664D03"
C_HELPER_BG  = "F5F5F5"
C_TBL_HDR    = "2C3E6B"
C_TBL_ALT    = "E8ECF4"
C_BORDER     = "C5CED8"
C_SUB_BG     = "E0E7F1"
C_SUB_TXT    = "2C3E6B"


# ================================================================
# STYLES
# ================================================================
F_TITLE       = Font(name="Arial", size=18, bold=True, color=C_NAVY)
F_SUBTITLE    = Font(name="Arial", size=10, color=C_BLUE)
F_INSTRUCT    = Font(name="Arial", size=9, italic=True, color=C_HINT)
F_SECTION     = Font(name="Arial", size=11, bold=True, color=C_WHITE)
F_SUBSECTION  = Font(name="Arial", size=9, bold=True, color=C_SUB_TXT)
F_LABEL       = Font(name="Arial", size=10, color=C_TEXT)
F_INPUT       = Font(name="Arial", size=10, bold=True, color=C_TEXT)
F_OUT_LABEL   = Font(name="Arial", size=10, bold=True, color=C_TEXT)
F_OUT_VALUE   = Font(name="Arial", size=11, bold=True, color=C_NAVY)
F_TOTAL_LABEL = Font(name="Arial", size=12, bold=True, color=C_NAVY)
F_TOTAL_VALUE = Font(name="Arial", size=14, bold=True, color=C_NAVY)
F_RNG_LABEL   = Font(name="Arial", size=10, color=C_TEXT)
F_RNG_VALUE   = Font(name="Arial", size=10, color=C_NAVY)
F_RNG_HDR     = Font(name="Arial", size=8, bold=True, color=C_HINT)
F_HINT        = Font(name="Arial", size=9, color=C_HINT, italic=True)
F_HELPER      = Font(name="Arial", size=9, color="A0A0A0")
F_TBL_HDR     = Font(name="Arial", size=10, bold=True, color=C_WHITE)
F_TBL_BODY    = Font(name="Arial", size=10, color=C_TEXT)
F_TBL_TITLE   = Font(name="Arial", size=12, bold=True, color=C_NAVY)
F_DISCL       = Font(name="Arial", size=8, color=C_DISCL_TXT)
F_SRC         = Font(name="Arial", size=8, color=C_HINT, italic=True)

P_SECTION   = PatternFill(start_color=C_MED_NAVY, end_color=C_MED_NAVY, fill_type="solid")
P_SUBSECT   = PatternFill(start_color=C_SUB_BG, end_color=C_SUB_BG, fill_type="solid")
P_INPUT     = PatternFill(start_color=C_INPUT_BG, end_color=C_INPUT_BG, fill_type="solid")
P_OUTPUT    = PatternFill(start_color=C_OUTPUT_BG, end_color=C_OUTPUT_BG, fill_type="solid")
P_TOTAL     = PatternFill(start_color=C_TOTAL_BG, end_color=C_TOTAL_BG, fill_type="solid")
P_RANGE     = PatternFill(start_color=C_RANGE_BG, end_color=C_RANGE_BG, fill_type="solid")
P_DISCL     = PatternFill(start_color=C_DISCL_BG, end_color=C_DISCL_BG, fill_type="solid")
P_HELPER    = PatternFill(start_color=C_HELPER_BG, end_color=C_HELPER_BG, fill_type="solid")
P_TBL_HDR   = PatternFill(start_color=C_TBL_HDR, end_color=C_TBL_HDR, fill_type="solid")
P_TBL_ALT   = PatternFill(start_color=C_TBL_ALT, end_color=C_TBL_ALT, fill_type="solid")

B_LIGHT = Border(
    left=Side(style="thin", color=C_BORDER),
    right=Side(style="thin", color=C_BORDER),
    top=Side(style="thin", color=C_BORDER),
    bottom=Side(style="thin", color=C_BORDER),
)
B_INPUT = Border(
    left=Side(style="thin", color=C_INPUT_BD),
    right=Side(style="thin", color=C_INPUT_BD),
    top=Side(style="thin", color=C_INPUT_BD),
    bottom=Side(style="thin", color=C_INPUT_BD),
)
B_RANGE = Border(
    left=Side(style="thin", color=C_RANGE_BD),
    right=Side(style="thin", color=C_RANGE_BD),
    top=Side(style="thin", color=C_RANGE_BD),
    bottom=Side(style="thin", color=C_RANGE_BD),
)
B_DISCL = Border(
    left=Side(style="thin", color=C_DISCL_BD),
    right=Side(style="thin", color=C_DISCL_BD),
    top=Side(style="thin", color=C_DISCL_BD),
    bottom=Side(style="thin", color=C_DISCL_BD),
)

AL_C  = Alignment(horizontal="center", vertical="center")
AL_R  = Alignment(horizontal="right", vertical="center")
AL_L  = Alignment(horizontal="left", vertical="center")
AL_LI = Alignment(horizontal="left", vertical="center", indent=1)
AL_LW = Alignment(horizontal="left", vertical="top", wrap_text=True)

FMT_CHF   = '"CHF "#,##0'
FMT_PCT   = '+0.0%;-0.0%'
FMT_RATE  = '0.0%'
FMT_NUM   = '#,##0'
FMT_DEC   = '0.00'
FMT_COEFF = '0.000'


# ================================================================
# HELPER
# ================================================================
def sc(ws, ref, value=None, font=None, fill=None, border=None,
       alignment=None, number_format=None):
    cell = ws[ref]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    return cell


def fill_row(ws, row, cols, fill, border=None):
    for c in cols:
        cell = ws[f"{c}{row}"]
        cell.fill = fill
        if border:
            cell.border = border


def sub_header(ws, row, text):
    """Einfügen einer Unter-Überschrift als visuelle Gruppierung."""
    for col in ["A", "B", "C"]:
        sc(ws, f"{col}{row}", fill=P_SUBSECT)
    sc(ws, f"A{row}", text, F_SUBSECTION, P_SUBSECT, alignment=AL_L)
    ws.row_dimensions[row].height = 18


# ================================================================
# TARIF-BLATT
# ================================================================
def populate_tarif(ws):
    ws.sheet_properties.tabColor = C_BLUE
    ws.sheet_view.showGridLines = False

    for col, w in [("A", 18), ("B", 18), ("C", 14), ("D", 10),
                   ("E", 4), ("F", 18), ("G", 18), ("H", 14), ("I", 10)]:
        ws.column_dimensions[col].width = w

    def write_table(start_row, col_offset, title, data):
        cols = [get_column_letter(col_offset + i) for i in range(4)]
        c0, c1, c2, c3 = cols

        ws.merge_cells(f"{c0}{start_row}:{c3}{start_row}")
        sc(ws, f"{c0}{start_row}", title, F_TBL_HDR, P_TBL_HDR, alignment=AL_C)
        for c in cols[1:]:
            sc(ws, f"{c}{start_row}", fill=P_TBL_HDR)

        hr = start_row + 1
        headers = ["Untergrenze", "Obergrenze", "Basisbetrag", "Satz"]
        for c, h in zip(cols, headers):
            sc(ws, f"{c}{hr}", h, F_TBL_HDR, P_TBL_HDR, B_LIGHT, AL_C)

        for i, (lo, hi, base, rate) in enumerate(data, start=start_row + 2):
            fill = P_TBL_ALT if (i - start_row - 2) % 2 == 1 else None
            sc(ws, f"{c0}{i}", lo, F_TBL_BODY, fill, B_LIGHT, number_format=FMT_NUM)
            sc(ws, f"{c1}{i}", hi, F_TBL_BODY, fill, B_LIGHT, number_format=FMT_NUM)
            sc(ws, f"{c2}{i}", base, F_TBL_BODY, fill, B_LIGHT, number_format=FMT_NUM)
            sc(ws, f"{c3}{i}", rate, F_TBL_BODY, fill, B_LIGHT, number_format=FMT_RATE)
            ws.row_dimensions[i].height = 18
        ws.row_dimensions[start_row].height = 22
        ws.row_dimensions[hr].height = 20

    # ---- ZH Gerichtskosten ----
    zh_cc = [
        (0,         1000,          0,    0.25),
        (1000,      5000,        250,    0.20),
        (5000,     20000,       1050,    0.14),
        (20000,    80000,       3150,    0.08),
        (80000,   300000,       7950,    0.04),
        (300000, 1000000,      16750,    0.02),
        (1000000, 10000000,    30750,    0.01),
        (10000000, 9999999999, 120750,   0.005),
    ]
    write_table(1, 1, "ZH Gerichtskosten (GebV OG ZH)", zh_cc)

    # ---- ZH Parteientschädigung ----
    zh_pc = [
        (0,          5000,        0,    0.25),
        (5000,      10000,     1250,    0.23),
        (10000,     20000,     2400,    0.15),
        (20000,     40000,     3900,    0.11),
        (40000,     80000,     6100,    0.09),
        (80000,    160000,     9700,    0.06),
        (160000,   300000,    14500,    0.035),
        (300000,   600000,    19400,    0.02),
        (600000,  1000000,    25400,    0.015),
        (1000000, 4000000,    31400,    0.01),
        (4000000, 10000000,   61400,    0.0075),
        (10000000, 9999999999, 106400,  0.005),
    ]
    write_table(1, 6, u"ZH Parteientsch\u00e4digung (AnwGebV ZH)", zh_pc)

    # ---- AG Gerichtskosten ----
    ag_start = 17
    ag_cc = [
        (0,         6500,      900,  0.11),
        (6501,     13000,     1160,  0.07),
        (13001,    52000,     1290,  0.06),
        (52001,   100000,      770,  0.07),
        (100001,  200000,     4270,  0.035),
        (200001,  400000,     6870,  0.022),
        (400001,  800000,     9670,  0.015),
        (800001, 1600000,    13670,  0.01),
        (1600001, 3300000,   21670,  0.005),
        (3300001, 9999999999, 28270, 0.003),
    ]
    write_table(ag_start, 1, "AG Gerichtskosten (GebT AG)", ag_cc)

    # ---- AG Parteientschädigung ----
    ag_pc = [
        (0,          6160,     1110,  0.22),
        (6161,      12300,     1230,  0.20),
        (12301,     24600,     1850,  0.15),
        (24601,     49300,     2590,  0.12),
        (49301,     98600,     4070,  0.09),
        (98601,    184800,     6530,  0.064),
        (184801,   369600,    10230,  0.044),
        (369601,   739200,    14300,  0.033),
        (739201,  1478400,    20240,  0.025),
        (1478401, 3080000,    29040,  0.019),
        (3080001, 6160000,    44440,  0.014),
        (6160001, 9999999999, 69080,  0.01),
    ]
    write_table(ag_start, 6, u"AG Parteientsch\u00e4digung (AnwT AG)", ag_pc)

    # ---- Berechnungshinweis ----
    nr = 32
    ws.merge_cells(f"A{nr}:I{nr}")
    sc(ws, f"A{nr}",
       u"ZH: Basisbetrag + (Streitwert \u2013 Untergrenze) \u00d7 Satz   |   "
       u"AG: Basisbetrag + Streitwert \u00d7 Satz",
       F_HINT)
    ws.merge_cells(f"A{nr+1}:I{nr+1}")
    sc(ws, f"A{nr+1}",
       u"Hinweis: Im ersten ZH-Band gilt ein Mindestbetrag von CHF 150 (GK) bzw. CHF 100 (PE).",
       F_HINT)


# ================================================================
# KOEFFIZIENTEN-BLATT
# ================================================================
def populate_koeffizienten(ws):
    ws.sheet_properties.tabColor = "548235"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    sc(ws, "A1", u"Modell 6 \u2014 Regressionskoeffizienten", F_TBL_TITLE)
    ws.row_dimensions[1].height = 28

    for col, val in [("A", "Variable"), ("B", "Gerichtskosten"), ("C", u"Parteientsch.")]:
        sc(ws, f"{col}2", val, F_TBL_HDR, P_TBL_HDR, B_LIGHT, AL_C)
    ws.row_dimensions[2].height = 22

    # Koeffizienten: (Name, CC, PC) — Zeilen 3–26
    coefficients = [
        ("Intercept",              0.927,   1.888),
        ("ln(Baseline)",          -0.111,  -0.170),
        ("D_AG",                  -0.518,  -3.414),
        ("D_AG x ln(Baseline)",    0.057,   0.463),
        (u"ln(Kl\u00e4ger)",       0.052,  -0.130),
        ("ln(Beklagte)",           0.092,   0.176),
        (u"Domizil Kl\u00e4ger",  -0.086,  -0.129),
        ("Domizil Beklagter",     -0.079,  -0.022),
        ("D_Bau",                  0.037,  -0.208),
        ("D_Handel",               0.007,   0.004),
        ("D_Finanz/Versicherung",  0.068,  -0.120),
        ("D_Organhaftung",        -0.006,  -0.064),
        ("D_UWG/URG",              0.592,   1.058),
        (u"D_Immaterialg\u00fcterrecht", 0.069, -0.079),
        ("D_Gesellschaftsrecht",  -0.227,  -0.045),
        (u"D_\u00dcbrige",         0.029,  -0.159),
        ("Klagantwort",           -0.088,  -0.004),
        ("Zweiter Schriftenwechsel", 0.504, 0.319),
        ("Widerklage",            -0.159,  -0.173),
        ("Widerklage 2. SW",       0.395,   0.192),
        (u"Zus\u00e4tzliche Eingaben", 0.051, 0.147),
        ("Vergleichsverhandlung",  0.150,   0.023),
        ("Gutachten",              0.123,  -0.319),
        (u"Unterliegensquote Kl\u00e4ger", 0.000, 0.122),
    ]

    for i, (name, cc, pc) in enumerate(coefficients, start=3):
        fill = P_TBL_ALT if (i - 3) % 2 == 1 else None
        sc(ws, f"A{i}", name, F_TBL_BODY, fill, B_LIGHT, AL_L)
        sc(ws, f"B{i}", cc, F_TBL_BODY, fill, B_LIGHT, AL_R, FMT_COEFF)
        sc(ws, f"C{i}", pc, F_TBL_BODY, fill, B_LIGHT, AL_R, FMT_COEFF)
        ws.row_dimensions[i].height = 18

    ws.row_dimensions[28].height = 10

    # ---- Rechtsgebiet-Zuordnung (Zeilen 29–39) ----
    ws.merge_cells("A29:C29")
    sc(ws, "A29", "Rechtsgebiet-Zuordnung",
       Font(name="Arial", size=11, bold=True, color=C_NAVY))
    ws.row_dimensions[29].height = 24

    for col, val in [("A", "Rechtsgebiet"), ("B", "Gerichtskosten"), ("C", u"Parteientsch.")]:
        sc(ws, f"{col}30", val, F_TBL_HDR, P_TBL_HDR, B_LIGHT, AL_C)
    ws.row_dimensions[30].height = 22

    # WICHTIG: Namen müssen exakt mit den Dropdown-Werten auf dem Rechner-Blatt übereinstimmen.
    rechtsgebiete = [
        ("Forderung Dienstleistungen",       0.000,   0.000),
        ("Forderung Bau",                    0.037,  -0.208),
        ("Forderung Handel",                 0.007,   0.004),
        ("Forderung Finanz/Versicherung",    0.068,  -0.120),
        ("Haftung Gesellschaftsorgane",     -0.006,  -0.064),
        ("UWG/URG",                          0.592,   1.058),
        (u"Immaterialg\u00fcterrecht",       0.069,  -0.079),
        ("Gesellschaftsrecht",              -0.227,  -0.045),
        (u"\u00dcbrige",                     0.029,  -0.159),
    ]

    for i, (name, cc, pc) in enumerate(rechtsgebiete, start=31):
        fill = P_TBL_ALT if (i - 31) % 2 == 1 else None
        sc(ws, f"A{i}", name, F_TBL_BODY, fill, B_LIGHT, AL_L)
        sc(ws, f"B{i}", cc, F_TBL_BODY, fill, B_LIGHT, AL_R, FMT_COEFF)
        sc(ws, f"C{i}", pc, F_TBL_BODY, fill, B_LIGHT, AL_R, FMT_COEFF)
        ws.row_dimensions[i].height = 18

    nr = 42
    ws.merge_cells(f"A{nr}:C{nr}")
    sc(ws, f"A{nr}",
       u"Quelle: Modell 6, Tabelle 3.5 (Year FE + Judge FE). "
       u"Medianer abs. Prognosefehler: GK 21%, PE 25.8%. "
       u"Referenzkategorie: Forderung Dienstleistungen.",
       F_SRC)


# ================================================================
# RECHNER-BLATT
# ================================================================
def build_rechner(ws):
    ws.sheet_properties.tabColor = "BF8F00"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 36

    # ---- TITEL ----
    ws.row_dimensions[1].height = 8
    ws.merge_cells("A2:C2")
    sc(ws, "A2", "Prozesskostenrechner", F_TITLE)
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:C3")
    sc(ws, "A3",
       u"Ordentliche Verfahren vor den Handelsgerichten Z\u00fcrich und Aargau \u2014 Modell 6",
       F_SUBTITLE)
    ws.row_dimensions[3].height = 18

    # ---- Bedienungshinweis ----
    ws.merge_cells("A4:C4")
    sc(ws, "A4",
       u"Bitte f\u00fcllen Sie die gelb markierten Felder aus. "
       u"Die Ergebnisse werden automatisch berechnet. "
       u"Das Tool sch\u00e4tzt Gerichtskosten (vom Gericht festgesetzt) und "
       u"Parteientsch\u00e4digung (Anwaltskostenersatz an die obsiegende Partei) "
       u"auf Basis von \u00fcber 900 Urteilen.",
       F_INSTRUCT, alignment=AL_LW)
    ws.row_dimensions[4].height = 36

    # ---- EINGABE-Header ----
    ws.merge_cells("A5:C5")
    sc(ws, "A5", "   EINGABE", F_SECTION, P_SECTION, alignment=AL_L)
    fill_row(ws, 5, ["B", "C"], P_SECTION)
    ws.row_dimensions[5].height = 26

    # ---- Unter-Überschrift: Fall ----
    sub_header(ws, 6, "  Fall")

    # ---- Eingabefelder ----
    # WICHTIG: Zeilen 7–21 für Eingaben. Formeln referenzieren diese Zellen!
    # B7=Kanton, B8=Streitwert, B9=Rechtsgebiet,
    # B10=DomKl, B11=DomBekl, B12=AnzKl, B13=AnzBekl,
    # B15=Klagantwort, B16=2.SW, B17=Widerklage, B18=WK2SW,
    # B19=Zusätzl, B20=Vergleich, B21=Gutachten, B22=Unterliegensquote

    inputs_fall = [
        (7,  "Kanton",                                u"Z\u00fcrich",
         u"Standort des zust\u00e4ndigen Handelsgerichts"),
        (8,  "Streitwert (CHF)",                      500000,
         u"Gesamtwert aller Rechtsbegehren"),
        (9,  "Rechtsgebiet",                          "Forderung Dienstleistungen",
         u"Hauptkategorie gem\u00e4ss Urteilskopf"),
        (10, u"Kl\u00e4ger mit Sitz in der Schweiz",  "Ja",
         u"Sitz oder Wohnsitz in der Schweiz?"),
        (11, "Beklagter mit Sitz in der Schweiz",     "Ja",
         u"Sitz oder Wohnsitz in der Schweiz?"),
        (12, u"Anzahl Kl\u00e4ger",                   1,
         u"inkl. allfälliger Streitgenossen"),
        (13, "Anzahl Beklagte",                       1,
         u"inkl. allfälliger Streitgenossen"),
    ]

    for row, label, default, hint in inputs_fall:
        sc(ws, f"A{row}", label, F_LABEL, alignment=AL_LI)
        cell = sc(ws, f"B{row}", default, F_INPUT, P_INPUT, B_INPUT, AL_C)
        if row == 8:
            cell.number_format = FMT_CHF
        if hint:
            sc(ws, f"C{row}", hint, F_HINT)
        ws.row_dimensions[row].height = 24

    # ---- Unter-Überschrift: Verfahrensgang ----
    sub_header(ws, 14, u"  Verfahrensgang (erwartete oder eingetretene Prozessschritte)")

    inputs_verf = [
        (15, "Klagantwort eingereicht",               "Ja",
         u"Hat der Beklagte eine Klagantwort eingereicht?"),
        (16, "Zweiter Schriftenwechsel",              "Nein",
         u"Replik und Duplik angeordnet?"),
        (17, "Widerklage erhoben",                    "Nein",
         u"Hat der Beklagte eine Widerklage erhoben?"),
        (18, "Widerklage mit 2. Schriftenwechsel",    "Nein",
         u"Zweiter Schriftenwechsel zur Widerklage?"),
        (19, u"Zus\u00e4tzliche Eingaben",            "Nein",
         u"Weitere Eingaben \u00fcber die Schriftenw. hinaus?"),
        (20, "Vergleichsverhandlung",                 "Nein",
         u"Wurde eine Vergleichsverhandlung durchgef\u00fchrt?"),
        (21, "Gutachten",                             "Nein",
         u"Wurde ein Sachverst\u00e4ndigengutachten eingeholt?"),
    ]

    for row, label, default, hint in inputs_verf:
        sc(ws, f"A{row}", label, F_LABEL, alignment=AL_LI)
        sc(ws, f"B{row}", default, F_INPUT, P_INPUT, B_INPUT, AL_C)
        if hint:
            sc(ws, f"C{row}", hint, F_HINT)
        ws.row_dimensions[row].height = 24

    # ---- Unter-Überschrift: Kostenverteilung ----
    sub_header(ws, 22, "  Kostenverteilung")

    sc(ws, "A23", u"Unterliegensquote Kl\u00e4ger (0\u20131)", F_LABEL, alignment=AL_LI)
    cell = sc(ws, "B23", 0.5, F_INPUT, P_INPUT, B_INPUT, AL_C)
    cell.number_format = FMT_DEC
    sc(ws, "C23",
       u"0 = Kl. obsiegt, 1 = Kl. unterliegt (beeinflusst nur PE)",
       F_HINT)
    ws.row_dimensions[23].height = 24

    # ---- DATENVALIDIERUNGEN ----
    dv_kanton = DataValidation(
        type="list", formula1=u'"Z\u00fcrich,Aargau"', allow_blank=False)
    dv_kanton.error = u"Bitte Z\u00fcrich oder Aargau w\u00e4hlen."
    dv_kanton.prompt = u"Kanton w\u00e4hlen"
    dv_kanton.showErrorMessage = True
    dv_kanton.showInputMessage = True
    ws.add_data_validation(dv_kanton)
    dv_kanton.add(ws["B7"])

    rg_list = (u'"Forderung Dienstleistungen,Forderung Bau,Forderung Handel,'
               u'Forderung Finanz/Versicherung,Haftung Gesellschaftsorgane,'
               u'UWG/URG,Immaterialg\u00fcterrecht,Gesellschaftsrecht,\u00dcbrige"')
    dv_rg = DataValidation(type="list", formula1=rg_list, allow_blank=False)
    dv_rg.error = "Bitte ein Rechtsgebiet aus der Liste ausw\u00e4hlen."
    dv_rg.prompt = u"Rechtsgebiet w\u00e4hlen"
    dv_rg.showErrorMessage = True
    dv_rg.showInputMessage = True
    ws.add_data_validation(dv_rg)
    dv_rg.add(ws["B9"])

    dv_jn = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=False)
    dv_jn.error = "Bitte Ja oder Nein ausw\u00e4hlen."
    dv_jn.prompt = u"Ja oder Nein w\u00e4hlen"
    dv_jn.showErrorMessage = True
    dv_jn.showInputMessage = True
    ws.add_data_validation(dv_jn)
    for r in [10, 11, 15, 16, 17, 18, 19, 20, 21]:
        dv_jn.add(ws[f"B{r}"])

    dv_count = DataValidation(type="whole", operator="greaterThanOrEqual",
                              formula1="1", allow_blank=False)
    dv_count.error = "Mindestens 1."
    dv_count.prompt = "Ganze Zahl eingeben (mindestens 1)"
    dv_count.showErrorMessage = True
    dv_count.showInputMessage = True
    ws.add_data_validation(dv_count)
    dv_count.add(ws["B12"])
    dv_count.add(ws["B13"])

    dv_loss = DataValidation(type="decimal", operator="between",
                             formula1="0", formula2="1", allow_blank=False)
    dv_loss.error = "Wert muss zwischen 0 und 1 liegen."
    dv_loss.prompt = (u"0 = Kl\u00e4ger obsiegt vollst\u00e4ndig, "
                      u"1 = Kl\u00e4ger unterliegt vollst\u00e4ndig")
    dv_loss.showErrorMessage = True
    dv_loss.showInputMessage = True
    ws.add_data_validation(dv_loss)
    dv_loss.add(ws["B23"])

    dv_dv = DataValidation(type="whole", operator="greaterThanOrEqual",
                           formula1="1", allow_blank=False)
    dv_dv.error = "Bitte positiven Streitwert eingeben."
    dv_dv.prompt = "Streitwert in CHF (ganze Zahl)"
    dv_dv.showErrorMessage = True
    dv_dv.showInputMessage = True
    ws.add_data_validation(dv_dv)
    dv_dv.add(ws["B8"])

    # ---- Abstand ----
    ws.row_dimensions[24].height = 10

    # ================================================================
    # ERGEBNISSE
    # ================================================================
    ws.merge_cells("A25:C25")
    sc(ws, "A25", "   ERGEBNISSE", F_SECTION, P_SECTION, alignment=AL_L)
    fill_row(ws, 25, ["B", "C"], P_SECTION)
    ws.row_dimensions[25].height = 26

    # ---- TARIFFORMELN ----
    # Tarif-Blatt:  ZH CC: A3:A10/C3:C10/D3:D10,  ZH PC: F3:F14/H3:H14/I3:I14
    #               AG CC: A19:A28/C19:C28/D19:D28, AG PC: F19:F30/H19:H30/I19:I30

    T = "Tarif"

    baseline_cc = (
        f'=IF(B7="Z\u00fcrich",'
        f'MAX(150,'
        f'INDEX({T}!C$3:C$10,MATCH(B8,{T}!A$3:A$10,1))'
        f'+(B8-INDEX({T}!A$3:A$10,MATCH(B8,{T}!A$3:A$10,1)))'
        f'*INDEX({T}!D$3:D$10,MATCH(B8,{T}!A$3:A$10,1))),'
        f'INDEX({T}!C$19:C$28,MATCH(B8,{T}!A$19:A$28,1))'
        f'+B8*INDEX({T}!D$19:D$28,MATCH(B8,{T}!A$19:A$28,1)))'
    )

    baseline_pc = (
        f'=IF(B7="Z\u00fcrich",'
        f'MAX(100,'
        f'INDEX({T}!H$3:H$14,MATCH(B8,{T}!F$3:F$14,1))'
        f'+(B8-INDEX({T}!F$3:F$14,MATCH(B8,{T}!F$3:F$14,1)))'
        f'*INDEX({T}!I$3:I$14,MATCH(B8,{T}!F$3:F$14,1))),'
        f'INDEX({T}!H$19:H$30,MATCH(B8,{T}!F$19:F$30,1))'
        f'+B8*INDEX({T}!I$19:I$30,MATCH(B8,{T}!F$19:F$30,1)))'
    )

    # Zeile 26: Tarif GK
    sc(ws, "A26", u"Gerichtskosten gem\u00e4ss Tarif",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B26", baseline_cc, F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_CHF)
    sc(ws, "C26",
       u"Gesetzlicher Ausgangswert (GebV OG ZH / GebT AG)",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[26].height = 24

    # Zeile 27: Tarif PE
    sc(ws, "A27", u"Parteientsch\u00e4digung gem\u00e4ss Tarif",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B27", baseline_pc, F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_CHF)
    sc(ws, "C27",
       u"Gesetzlicher Ausgangswert (AnwGebV ZH / AnwT AG)",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[27].height = 24

    ws.row_dimensions[28].height = 6

    # ---- EMPIRISCHE SCHÄTZUNG ----
    K = "Koeffizienten"

    ln_dev_cc = (
        f'={K}!B3'
        f'+LN(B26)*{K}!B4'
        f'+IF(B7="Aargau",1,0)*{K}!B5'
        f'+IF(B7="Aargau",1,0)*LN(B26)*{K}!B6'
        f'+INDEX({K}!B$31:B$39,MATCH(B9,{K}!A$31:A$39,0))'
        f'+IF(B10="Ja",1,0)*{K}!B9'
        f'+IF(B11="Ja",1,0)*{K}!B10'
        f'+LN(MAX(1,B12))*{K}!B7'
        f'+LN(MAX(1,B13))*{K}!B8'
        f'+IF(B15="Ja",1,0)*{K}!B19'
        f'+IF(B16="Ja",1,0)*{K}!B20'
        f'+IF(B17="Ja",1,0)*{K}!B21'
        f'+IF(B18="Ja",1,0)*{K}!B22'
        f'+IF(B19="Ja",1,0)*{K}!B23'
        f'+IF(B20="Ja",1,0)*{K}!B24'
        f'+IF(B21="Ja",1,0)*{K}!B25'
    )

    ln_dev_pc = (
        f'={K}!C3'
        f'+LN(B27)*{K}!C4'
        f'+IF(B7="Aargau",1,0)*{K}!C5'
        f'+IF(B7="Aargau",1,0)*LN(B27)*{K}!C6'
        f'+INDEX({K}!C$31:C$39,MATCH(B9,{K}!A$31:A$39,0))'
        f'+IF(B10="Ja",1,0)*{K}!C9'
        f'+IF(B11="Ja",1,0)*{K}!C10'
        f'+LN(MAX(1,B12))*{K}!C7'
        f'+LN(MAX(1,B13))*{K}!C8'
        f'+IF(B15="Ja",1,0)*{K}!C19'
        f'+IF(B16="Ja",1,0)*{K}!C20'
        f'+IF(B17="Ja",1,0)*{K}!C21'
        f'+IF(B18="Ja",1,0)*{K}!C22'
        f'+IF(B19="Ja",1,0)*{K}!C23'
        f'+IF(B20="Ja",1,0)*{K}!C24'
        f'+IF(B21="Ja",1,0)*{K}!C25'
        f'+B23*{K}!C26'
    )

    # Hilfsberechnungen (Zeilen 46–47)
    sc(ws, "A46", "HILFSBERECHNUNGEN", F_HELPER, P_HELPER)
    sc(ws, "B46", None, fill=P_HELPER)
    sc(ws, "C46", None, fill=P_HELPER)
    ws.row_dimensions[46].height = 16

    sc(ws, "A47", "ln(GK / Grundtarif)", F_HELPER, P_HELPER, B_LIGHT)
    sc(ws, "B47", ln_dev_cc, F_HELPER, P_HELPER, B_LIGHT, number_format="0.0000")
    sc(ws, "C47", None, fill=P_HELPER)

    sc(ws, "A48", "ln(PE / Grundtarif)", F_HELPER, P_HELPER, B_LIGHT)
    sc(ws, "B48", ln_dev_pc, F_HELPER, P_HELPER, B_LIGHT, number_format="0.0000")
    sc(ws, "C48", None, fill=P_HELPER)

    # Zeile 29: Geschätzte GK
    sc(ws, "A29", u"Gesch\u00e4tzte Gerichtskosten",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B29", "=B26*EXP(B47)", F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_CHF)
    sc(ws, "C29",
       u"Modellsch\u00e4tzung basierend auf 976 Urteilen",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[29].height = 24

    # Zeile 30: Geschätzte PE
    sc(ws, "A30", u"Gesch\u00e4tzte Parteientsch\u00e4digung",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B30", "=B27*EXP(B48)", F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_CHF)
    sc(ws, "C30",
       u"Modellsch\u00e4tzung basierend auf 898 Urteilen",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[30].height = 24

    ws.row_dimensions[31].height = 4

    # Zeile 32: Gesamtkosten
    sc(ws, "A32", u"GESCH\u00c4TZTE GESAMTKOSTEN",
       F_TOTAL_LABEL, P_TOTAL, B_LIGHT, AL_L)
    sc(ws, "B32", "=B29+B30", F_TOTAL_VALUE, P_TOTAL, B_LIGHT, AL_R, FMT_CHF)
    sc(ws, "C32",
       u"Summe GK + PE (ohne eigene Anwaltskosten)",
       F_HINT, P_TOTAL, B_LIGHT)
    ws.row_dimensions[32].height = 30

    # ---- Bandbreite ----
    ws.merge_cells("A33:C33")
    sc(ws, "A33",
       u"Bandbreite: In 50% der F\u00e4lle liegen die tats\u00e4chlichen "
       u"Kosten innerhalb dieses Bereichs.",
       Font(name="Arial", size=8, italic=True, color=C_HINT))
    ws.row_dimensions[33].height = 16

    sc(ws, "A34", u"Bandbreite Gerichtskosten (\u00b121%)",
       F_RNG_LABEL, P_RANGE, B_RANGE, AL_L)
    sc(ws, "B34", "=B29*(1-0.21)", F_RNG_VALUE, P_RANGE, B_RANGE, AL_R, FMT_CHF)
    sc(ws, "C34", "=B29*(1+0.21)", F_RNG_VALUE, P_RANGE, B_RANGE, AL_R, FMT_CHF)
    ws.row_dimensions[34].height = 26

    sc(ws, "A35", u"Bandbreite Parteientsch\u00e4digung (\u00b125.8%)",
       F_RNG_LABEL, P_RANGE, B_RANGE, AL_L)
    sc(ws, "B35", "=B30*(1-0.258)", F_RNG_VALUE, P_RANGE, B_RANGE, AL_R, FMT_CHF)
    sc(ws, "C35", "=B30*(1+0.258)", F_RNG_VALUE, P_RANGE, B_RANGE, AL_R, FMT_CHF)
    ws.row_dimensions[35].height = 26

    ws.row_dimensions[36].height = 6

    # ---- Abweichung ----
    sc(ws, "A37", u"Abweichung vom Tarif \u2014 Gerichtskosten",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B37", "=(B29-B26)/B26", F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_PCT)
    sc(ws, "C37",
       u"positiv = \u00fcber Tarif, negativ = unter Tarif",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[37].height = 24

    sc(ws, "A38", u"Abweichung vom Tarif \u2014 Parteientsch\u00e4digung",
       F_OUT_LABEL, P_OUTPUT, B_LIGHT, AL_L)
    sc(ws, "B38", "=(B30-B27)/B27", F_OUT_VALUE, P_OUTPUT, B_LIGHT, AL_R, FMT_PCT)
    sc(ws, "C38",
       u"positiv = \u00fcber Tarif, negativ = unter Tarif",
       F_HINT, P_OUTPUT, B_LIGHT)
    ws.row_dimensions[38].height = 24

    # ================================================================
    # HAFTUNGSAUSSCHLUSS
    # ================================================================
    ws.row_dimensions[39].height = 10

    disclaimer = (
        u"HAFTUNGSAUSSCHLUSS: Dieses Tool dient ausschliesslich zu Informations- "
        u"und Bildungszwecken. Die Kostensch\u00e4tzungen basieren auf statistischen "
        u"Regressionsmodellen und empirischen Daten gem\u00e4ss der Studie "
        u"\u00abFrom Tariff to Award\u00bb. Die Ergebnisse sind N\u00e4herungswerte "
        u"und k\u00f6nnen erheblich von den tats\u00e4chlich zugesprochenen "
        u"Prozesskosten abweichen.\n\n"
        u"Der Autor \u00fcbernimmt keinerlei Haftung f\u00fcr die Richtigkeit, "
        u"Vollst\u00e4ndigkeit, Aktualit\u00e4t oder Zuverl\u00e4ssigkeit der mit "
        u"diesem Tool generierten Sch\u00e4tzungen. Die Nutzung dieses Tools stellt "
        u"keine Rechtsberatung dar. Es werden keinerlei Gew\u00e4hrleistungen "
        u"abgegeben. F\u00fcr fallspezifische Kosteneinsch\u00e4tzungen ist eine "
        u"unabh\u00e4ngige fachkundige Beratung einzuholen."
    )

    ws.merge_cells("A40:C44")
    sc(ws, "A40", disclaimer, F_DISCL, P_DISCL, B_DISCL, AL_LW)
    for r in range(40, 45):
        for c in ["A", "B", "C"]:
            ws[f"{c}{r}"].fill = P_DISCL
            ws[f"{c}{r}"].border = B_DISCL
        ws.row_dimensions[r].height = 18

    ws.merge_cells("A45:C45")
    sc(ws, "A45",
       u"Grundlage: \u00abFrom Tariff to Award\u00bb \u2014 Modell 6 "
       u"(Year FE + Judge FE). Medianer abs. Prognosefehler: GK 21%, PE 25.8%.",
       F_SRC)
    ws.row_dimensions[45].height = 16


# ================================================================
# MAIN
# ================================================================
def main():
    wb = openpyxl.Workbook()

    ws_rechner = wb.active
    ws_rechner.title = "Rechner"

    ws_tarif = wb.create_sheet("Tarif")
    ws_koeff = wb.create_sheet("Koeffizienten")

    populate_tarif(ws_tarif)
    populate_koeffizienten(ws_koeff)
    build_rechner(ws_rechner)

    wb.move_sheet("Rechner", offset=-2)

    path = "/Users/adriankonig/Desktop/Prozesskostenrechner.xlsx"
    wb.save(path)
    print(f"Gespeichert: {path}")


if __name__ == "__main__":
    main()
